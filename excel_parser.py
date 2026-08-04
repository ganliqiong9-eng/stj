#!/usr/bin/env python3
"""Parse Excel file and output SQL CREATE + INSERT statements for SQLite.

Usage: python3 excel_parser.py <xlsx_path> [--table-name NAME] [--db-path DB]
       python3 excel_parser.py <xlsx_path> --read-schema-only

Output: JSON with table_name, columns, rows, sql, error
"""

import json, sys, os, re, sqlite3
from openpyxl import load_workbook

def clean_name(name):
    """Clean a string to be a valid SQLite identifier."""
    name = re.sub(r'[^a-zA-Z0-9_\u4e00-\u9fff]', '_', str(name))
    if not name or name[0].isdigit():
        name = 'col_' + name
    return name

def infer_type(values):
    """Infer SQLite column type from sample values."""
    int_count = 0
    float_count = 0
    text_count = 0
    for v in values:
        if v is None:
            continue
        if isinstance(v, int):
            int_count += 1
        elif isinstance(v, float):
            float_count += 1
        elif isinstance(v, str):
            text_count += 1
    if int_count > 0 and float_count == 0 and text_count == 0:
        return 'INTEGER'
    if float_count > 0 or int_count > 0:
        return 'REAL'
    return 'TEXT'

def parse_excel(filepath, table_name=None):
    """Parse Excel file and return JSON with metadata and data."""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return {'error': 'Empty spreadsheet'}

        # First row as headers
        headers = [str(h) if h is not None else f'col_{i}' for i, h in enumerate(rows[0])]
        clean_headers = [clean_name(h) for h in headers]

        # Data rows
        data_rows = []
        for row in rows[1:]:
            processed = []
            for cell in row:
                if cell is None:
                    processed.append(None)
                elif isinstance(cell, (int, float)):
                    processed.append(cell)
                else:
                    processed.append(str(cell))
            data_rows.append(processed)

        # Remove empty trailing rows
        while data_rows and all(v is None for v in data_rows[-1]):
            data_rows.pop()

        if not data_rows:
            return {'error': 'No data rows found'}

        if not table_name:
            table_name = os.path.splitext(os.path.basename(filepath))[0]
            table_name = clean_name(table_name)

        # Infer types
        types = []
        for i in range(len(clean_headers)):
            col_values = [row[i] for row in data_rows if i < len(row)]
            types.append(infer_type(col_values))

        # Generate CREATE TABLE SQL
        col_defs = []
        for i in range(len(clean_headers)):
            col_defs.append(f'"{clean_headers[i]}" {types[i]}')
        sep = ",\n    "
        create_sql = f'CREATE TABLE IF NOT EXISTS "{table_name}" (\n    {sep.join(col_defs)}\n);'

        # Generate INSERT SQL
        insert_sql_parts = []
        for row in data_rows:
            vals = []
            for i in range(len(clean_headers)):
                v = row[i] if i < len(row) else None
                if v is None:
                    vals.append('NULL')
                elif isinstance(v, int):
                    vals.append(str(v))
                elif isinstance(v, float):
                    vals.append(str(v))
                else:
                    escaped = str(v).replace("'", "''")
                    vals.append(f"'{escaped}'")
            if vals:
                insert_sql_parts.append(f'INSERT INTO "{table_name}" VALUES ({", ".join(vals)});')

        return {
            'table_name': table_name,
            'columns': [{'name': clean_headers[i], 'type': types[i], 'original': headers[i]} for i in range(len(clean_headers))],
            'rows': len(data_rows),
            'create_sql': create_sql,
            'insert_sql': '\n'.join(insert_sql_parts)
        }
    except Exception as e:
        return {'error': str(e)}

if __name__ == '__main__':
    args = sys.argv[1:]
    if not args:
        print(json.dumps({'error': 'Usage: excel_parser.py <xlsx_path> [--table-name NAME]'}))
        sys.exit(1)

    filepath = args[0]
    table_name = None

    if '--table-name' in args:
        idx = args.index('--table-name')
        if idx + 1 < len(args):
            table_name = args[idx + 1]

    result = parse_excel(filepath, table_name)
    print(json.dumps(result, ensure_ascii=False))
