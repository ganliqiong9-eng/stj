#!/usr/bin/env python3
"""Analyze Excel files: field detection, type inference, description generation, folder suggestion.

Usage: python3 excel_analyzer.py <xlsx_path>
Output: JSON with table info, columns (name/type/description/samples), suggestedCategory
"""

import json, sys, os, re
from openpyxl import load_workbook

# Category detection keywords
CATEGORY_KEYWORDS = {
    '教育/学习': ['student', 'course', 'score', 'grade', 'class', 'teacher', 'exam', '学生', '课程', '成绩', '班级', '老师', '考试', 'score', 'enrollment', '选课', '学分'],
    '人力资源': ['employee', 'staff', 'department', 'salary', 'hire', 'name', '岗位', '员工', '部门', '薪资', '入职', '招聘', '考勤', 'attendance', 'leave'],
    '销售/客户': ['customer', 'order', 'product', 'price', 'amount', 'sales', '客户', '订单', '产品', '价格', '金额', '销售', 'contract', 'invoice'],
    '财务/会计': ['account', 'budget', 'revenue', 'cost', 'profit', 'expense', '财务', '预算', '收入', '成本', '利润', '支出', 'invoice', 'tax', '报表'],
    '项目管理': ['project', 'task', 'milestone', 'deadline', '进度', '项目', '任务', '里程碑', '截止', 'assignee', 'status', '状态', '负责人'],
    '库存/物流': ['inventory', 'stock', 'warehouse', 'supplier', '物流', '库存', '仓库', '供应商', 'shipment', 'delivery', '发货', '入库'],
    '技术/IT': ['server', 'database', 'api', 'code', 'config', '技术', '服务器', '数据库', '配置', 'deploy', 'version', '环境'],
}

def clean_name(name):
    name = re.sub(r'[^a-zA-Z0-9_\u4e00-\u9fff]', '_', str(name))
    if not name or name[0].isdigit():
        name = 'col_' + name
    return name

def infer_type(values):
    """Enhanced type inference with date detection."""
    if not values:
        return 'TEXT'
    non_null = [v for v in values if v is not None]
    if not non_null:
        return 'TEXT'

    # Check for boolean
    if all(isinstance(v, bool) or str(v).lower() in ('true', 'false', '是', '否', 'yes', 'no', '0', '1') for v in non_null):
        return 'BOOLEAN'

    # Check for date
    date_pattern = re.compile(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}([\sT]\d{1,2}:\d{2})?$')
    date_count = sum(1 for v in non_null if isinstance(v, str) and date_pattern.match(v))
    if date_count > len(non_null) * 0.6:
        return 'DATE'

    # Check numeric
    int_count = sum(1 for v in non_null if isinstance(v, int))
    float_count = sum(1 for v in non_null if isinstance(v, float))

    if int_count == len(non_null):
        return 'INTEGER'
    if float_count > 0 or int_count > 0:
        return 'REAL'
    return 'TEXT'

def generate_description(col_name, col_type, sample_values):
    """Generate human-readable field description."""
    name_cn = col_name
    patterns = {
        'id': '唯一标识',
        'name|姓名|名称': '名称/姓名',
        'email|邮件': '电子邮件地址',
        'phone|电话|mobile|手机': '联系电话',
        'address|地址': '地址信息',
        'city|城市': '所在城市',
        'date|时间|time|日期': '日期/时间',
        'price|金额|amount|cost|费用': '金额',
        'quantity|number|count|数量': '数量',
        'desc|描述|备注|remark|note|说明': '描述/备注',
        'status|状态': '状态',
        'type|类型|category|类别': '类型/类别',
        'score|成绩|分数': '分数/成绩',
        'percent|率|ratio|比例': '比例/百分比',
        'code|编码|代码': '编码',
        'url|link|链接|网址': '链接/网址',
    }

    desc = None
    for pattern, label in patterns.items():
        if re.search(pattern, col_name, re.IGNORECASE):
            desc = label
            break

    if not desc:
        desc = f'{col_name}'

    # Add type hint
    if col_type == 'INTEGER':
        desc += '（整数）'
    elif col_type == 'REAL':
        desc += '（数值）'
    elif col_type == 'DATE':
        desc += '（日期）'
    elif col_type == 'BOOLEAN':
        desc += '（布尔值）'

    # Add sample
    if sample_values:
        samples = [str(v) for v in sample_values[:3] if v is not None]
        if samples:
            desc += f' 例如: {", ".join(samples)}'

    return desc

def suggest_category(columns, sample_rows):
    """Auto-determine folder/category based on column names and data."""
    scores = {}
    all_names = ' '.join(c['name'].lower() for c in columns)
    all_values = ' '.join(str(v).lower() for row in sample_rows[:5] for v in row)

    for category, keywords in CATEGORY_KEYWORDS.items():
        score = 0
        for kw in keywords:
            score += all_names.count(kw.lower()) * 2
            score += all_values.count(kw.lower())
        if score > 0:
            scores[category] = score

    if scores:
        return max(scores, key=scores.get)
    return '通用'

def analyze_excel(filepath):
    """Analyze Excel file and return structured field information."""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return {'error': 'Empty spreadsheet'}

        headers = [str(h) if h is not None else f'col_{i}' for i, h in enumerate(rows[0])]
        clean_headers = [clean_name(h) for h in headers]

        data_rows = []
        for row in rows[1:]:
            processed = []
            for cell in row:
                if cell is None:
                    processed.append(None)
                elif isinstance(cell, (int, float)):
                    processed.append(cell)
                elif isinstance(cell, bool):
                    processed.append(cell)
                else:
                    processed.append(str(cell))
            data_rows.append(processed)

        while data_rows and all(v is None for v in data_rows[-1]):
            data_rows.pop()

        table_name = os.path.splitext(os.path.basename(filepath))[0]
        table_name = clean_name(table_name)

        columns_info = []
        for i in range(len(clean_headers)):
            col_values = [row[i] for row in data_rows if i < len(row) and row[i] is not None]
            col_type = infer_type(col_values)
            sample_values = col_values[:5]
            description = generate_description(clean_headers[i], col_type, sample_values)

            columns_info.append({
                'name': clean_headers[i],
                'originalName': headers[i],
                'type': col_type,
                'description': description,
                'sampleValues': [str(v) for v in sample_values[:3] if v is not None],
                'nonNullCount': len(col_values),
                'totalCount': len(data_rows),
            })

        suggested_category = suggest_category(columns_info, data_rows[:10])

        # Generate CREATE TABLE SQL
        type_map = {
            'INTEGER': 'INTEGER',
            'REAL': 'REAL',
            'BOOLEAN': 'INTEGER',
            'DATE': 'TEXT',
            'TEXT': 'TEXT',
        }
        col_defs = []
        for c in columns_info:
            sql_type = type_map.get(c['type'], 'TEXT')
            col_defs.append(f'    "{c["name"]}" {sql_type} -- {c["description"]}')
        create_sql = f'CREATE TABLE "{table_name}" (\n' + ',\n'.join(col_defs) + '\n);'

        return {
            'tableName': table_name,
            'originalFileName': os.path.basename(filepath),
            'rowCount': len(data_rows),
            'columnCount': len(columns_info),
            'columns': columns_info,
            'suggestedCategory': suggested_category,
            'createSql': create_sql,
        }
    except Exception as e:
        return {'error': str(e)}

if __name__ == '__main__':
    args = sys.argv[1:]
    if not args:
        print(json.dumps({'error': 'Usage: excel_analyzer.py <xlsx_path>'}))
        sys.exit(1)
    result = analyze_excel(args[0])
    print(json.dumps(result, ensure_ascii=False, indent=2))
